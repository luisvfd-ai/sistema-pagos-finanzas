from decimal import Decimal
import json
import csv
import io
import re
from datetime import datetime, date, timedelta
import os
from functools import wraps

from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMultiAlternatives
from django.db import transaction
from django.db.models import Sum, Value, DecimalField, Count, Q
from django.db.models.functions import Coalesce
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.utils.html import strip_tags
from django.template.loader import render_to_string
from django.conf import settings
from django.http import FileResponse

from .models import (
    PagoProgramado,
    PagoReal,
    MovimientoBancario,
    ImportacionPago,
    ImportacionPagoDetalle,
)
from .forms import (
    PagoProgramadoForm,
    PagoRealForm,
    CartolaImportForm,
    AutoConciliacionForm,
    PagosImportExcelForm,
)

from .dashboard import (
    obtener_kpis_financieros,
    flujo_proyectado_mensual_chart,
    calcular_riesgo_financiero,
    eventos_vencidos,
    eventos_proximos,
    resumen_alertas_financieras,
    obtener_panel_alertas_financieras,
    obtener_alertas_urgentes_email,
    resumen_alertas_urgentes_email,
    listar_compromisos_financieros,
    resumen_estados_compromisos,
)


VIEW_PERMISSION_MAP = {
    'dashboard_financiero': ['pagos.view_pagoprogramado'],
    'alertas_financieras': ['pagos.view_pagoprogramado'],
    'enviar_alerta_urgente_email': ['pagos.change_pagoprogramado'],
    'pagos_lista': ['pagos.view_pagoprogramado'],
    'pagos_crear': ['pagos.add_pagoprogramado'],
    'pagos_editar': ['pagos.change_pagoprogramado'],

    'pagos_importar_excel': ['pagos.add_importacionpago'],
    'pagos_importar_excel_limpiar_preview': ['pagos.add_importacionpago'],
    'pagos_importar_excel_confirmar': ['pagos.add_importacionpago'],
    'importaciones_historial': ['pagos.view_importacionpago'],
    'importacion_revertir': ['pagos.delete_importacionpago'],

    'pagos_real_crear': ['pagos.add_pagoreal'],
    'pagos_real_editar': ['pagos.change_pagoreal'],
    'reportes_financieros': ['pagos.view_pagoreal'],

    'cartolas_importar': ['pagos.add_movimientobancario'],
    'cartolas_lista': ['pagos.add_movimientobancario'],
    'cartolas_sugerencias': ['pagos.add_movimientobancario'],
    'cartolas_auto_conciliar': ['pagos.change_movimientobancario'],
    'conciliacion_panel': ['pagos.add_movimientobancario'],
    'cartolas_conciliar': ['pagos.change_movimientobancario'],
    'cartolas_desconciliar': ['pagos.change_movimientobancario'],

    'descargar_plantilla_importacion': ['pagos.add_importacionpago'],
    'descargar_guia_importacion': ['pagos.add_importacionpago'],
}


def staff_member_required(view_func):
    @wraps(view_func)
    @login_required
    def _wrapped_view(request, *args, **kwargs):
        if request.user.is_superuser:
            return view_func(request, *args, **kwargs)

        required_perms = VIEW_PERMISSION_MAP.get(view_func.__name__, [])

        if required_perms:
            if not request.user.has_perms(required_perms):
                messages.error(request, 'No tienes permisos para acceder a esta sección.')

                if request.user.has_perm('pagos.view_pagoprogramado'):
                    return redirect('dashboard_financiero')

                return redirect('logout_view')

        return view_func(request, *args, **kwargs)

    return _wrapped_view


# ==================================================
# FORM FILTRO REPORTES (para reportes.html)
# ==================================================

class ReportesFiltroForm(forms.Form):
    fecha_desde = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={"type": "date", "class": "form-control"})
    )
    fecha_hasta = forms.DateField(
        required=False,
        widget=forms.DateInput(attrs={"type": "date", "class": "form-control"})
    )


# ==================================================
# HELPERS EXPORT (reportes)
# ==================================================

def _get_rango_fechas_from_request(request):
    hoy = timezone.now().date()
    desde_default = hoy.replace(day=1)
    hasta_default = hoy

    if request.method == 'POST':
        form = ReportesFiltroForm(request.POST)
        if form.is_valid():
            desde = form.cleaned_data.get('fecha_desde') or desde_default
            hasta = form.cleaned_data.get('fecha_hasta') or hasta_default
        else:
            desde, hasta = desde_default, hasta_default
        return desde, hasta

    desde_raw = request.GET.get('fecha_desde')
    hasta_raw = request.GET.get('fecha_hasta')

    desde = parse_date(desde_raw) if desde_raw else None
    hasta = parse_date(hasta_raw) if hasta_raw else None

    if not desde:
        desde = desde_default
    if not hasta:
        hasta = hasta_default

    return desde, hasta


def _build_report_queryset(desde, hasta):
    return (
        PagoReal.objects.filter(fecha_pago__range=[desde, hasta])
        .select_related('pago')
        .order_by('fecha_pago', 'id')
    )


def _export_csv(pagos_qs, desde, hasta):
    filename = f"reporte_pagos_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.csv"

    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    response.write('\ufeff')

    writer = csv.writer(response)
    writer.writerow(["Fecha", "Compromiso", "Monto", "Método", "Observación"])

    for p in pagos_qs:
        writer.writerow([
            p.fecha_pago.strftime('%Y-%m-%d') if p.fecha_pago else "",
            p.pago.nombre if p.pago_id else "",
            str(p.monto),
            p.metodo_pago or "",
            (p.observacion or "").strip(),
        ])

    return response


def _export_xlsx(pagos_qs, desde, hasta):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse

    filename = f"reporte_pagos_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"

    headers = ["Fecha", "Compromiso", "Monto", "Método", "Observación"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for p in pagos_qs:
        ws.append([
            p.fecha_pago.strftime('%Y-%m-%d') if p.fecha_pago else "",
            p.pago.nombre if p.pago_id else "",
            float(p.monto or 0),
            p.metodo_pago or "",
            (p.observacion or "").strip(),
        ])

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '#,##0.00'

    widths = [12, 30, 14, 16, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_pdf(pagos_qs, desde, hasta, total, promedio):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse

    filename = f"reporte_pagos_{desde.strftime('%Y%m%d')}_{hasta.strftime('%Y%m%d')}.pdf"

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24,
        topMargin=24, bottomMargin=24
    )

    styles = getSampleStyleSheet()
    story = []

    title = Paragraph(
        f"<b>Reporte de pagos reales</b> ({desde.strftime('%d-%m-%Y')} → {hasta.strftime('%d-%m-%Y')})",
        styles["Title"]
    )
    story.append(title)
    story.append(Spacer(1, 10))

    resumen = Paragraph(
        f"<b>Total período:</b> ${total:,.0f} &nbsp;&nbsp; "
        f"<b>Cantidad:</b> {pagos_qs.count()} &nbsp;&nbsp; "
        f"<b>Promedio:</b> ${promedio:,.0f}",
        styles["Normal"]
    )
    story.append(resumen)
    story.append(Spacer(1, 12))

    data = [["Fecha", "Compromiso", "Monto", "Método", "Observación"]]
    for p in pagos_qs:
        data.append([
            p.fecha_pago.strftime('%Y-%m-%d') if p.fecha_pago else "",
            p.pago.nombre if p.pago_id else "",
            f"${(p.monto or 0):,.0f}",
            p.metodo_pago or "",
            (p.observacion or "—"),
        ])

    table = Table(data, colWidths=[75, 220, 90, 110, 280])
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("ALIGN", (2, 1), (2, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
    ]))

    story.append(table)
    doc.build(story)

    pdf = buffer.getvalue()
    buffer.close()

    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    response.write(pdf)
    return response


# ==================================================
# HELPERS ALERTAS / PRIORIDAD VISUAL
# ==================================================

def _get_alerta_fecha(item):
    if isinstance(item, dict):
        return item.get('fecha') or item.get('fecha_alerta')
    return getattr(item, 'fecha', None) or getattr(item, 'fecha_alerta', None)


def _build_alerta_prioridad(item):
    hoy = timezone.now().date()
    fecha_evento = _get_alerta_fecha(item)
    delta = (fecha_evento - hoy).days

    if delta < 0:
        return {
            'prioridad_label': 'VENCIDA',
            'prioridad_clase': 'danger',
            'prioridad_texto': 'text-danger',
            'prioridad_badge_bg': '#dc2626',
            'prioridad_badge_text': '#ffffff',
            'prioridad_fila_bg': '#fef2f2',
            'prioridad_fila_text': '#991b1b',
            'prioridad_orden': 1,
        }

    if delta == 0:
        return {
            'prioridad_label': 'HOY',
            'prioridad_clase': 'warning',
            'prioridad_texto': 'text-warning',
            'prioridad_badge_bg': '#f59e0b',
            'prioridad_badge_text': '#111827',
            'prioridad_fila_bg': '#fffbeb',
            'prioridad_fila_text': '#92400e',
            'prioridad_orden': 2,
        }

    if delta == 1:
        return {
            'prioridad_label': '+1 día',
            'prioridad_clase': 'warning',
            'prioridad_texto': 'text-warning',
            'prioridad_badge_bg': '#fbbf24',
            'prioridad_badge_text': '#111827',
            'prioridad_fila_bg': '#fefce8',
            'prioridad_fila_text': '#854d0e',
            'prioridad_orden': 3,
        }

    if delta == 2:
        return {
            'prioridad_label': '+2 días',
            'prioridad_clase': 'warning',
            'prioridad_texto': 'text-warning',
            'prioridad_badge_bg': '#fde68a',
            'prioridad_badge_text': '#111827',
            'prioridad_fila_bg': '#fefce8',
            'prioridad_fila_text': '#854d0e',
            'prioridad_orden': 4,
        }

    if delta == 3:
        return {
            'prioridad_label': '+3 días',
            'prioridad_clase': 'warning',
            'prioridad_texto': 'text-warning',
            'prioridad_badge_bg': '#fde68a',
            'prioridad_badge_text': '#111827',
            'prioridad_fila_bg': '#fefce8',
            'prioridad_fila_text': '#854d0e',
            'prioridad_orden': 5,
        }

    return {
        'prioridad_label': '4 a 7 días',
        'prioridad_clase': 'primary',
        'prioridad_texto': 'text-primary',
        'prioridad_badge_bg': '#2563eb',
        'prioridad_badge_text': '#ffffff',
        'prioridad_fila_bg': '#eff6ff',
        'prioridad_fila_text': '#1d4ed8',
        'prioridad_orden': 6,
    }


def _enriquecer_alertas_eventos(eventos):
    eventos_enriquecidos = []
    for e in eventos:
        prioridad = _build_alerta_prioridad(e)

        if isinstance(e, dict):
            item = dict(e)
            item.update(prioridad)
            eventos_enriquecidos.append(item)
            continue

        e.prioridad_label = prioridad['prioridad_label']
        e.prioridad_clase = prioridad['prioridad_clase']
        e.prioridad_texto = prioridad['prioridad_texto']
        e.prioridad_badge_bg = prioridad['prioridad_badge_bg']
        e.prioridad_badge_text = prioridad['prioridad_badge_text']
        e.prioridad_fila_bg = prioridad['prioridad_fila_bg']
        e.prioridad_fila_text = prioridad['prioridad_fila_text']
        e.prioridad_orden = prioridad['prioridad_orden']
        eventos_enriquecidos.append(e)
    return eventos_enriquecidos


def _enriquecer_panel_alertas(panel):
    return {
        'vencidas': _enriquecer_alertas_eventos(panel.get('vencidas', [])),
        'vencen_hoy': _enriquecer_alertas_eventos(panel.get('vencen_hoy', [])),
        'urgentes': _enriquecer_alertas_eventos(panel.get('urgentes', [])),
        'proximas': _enriquecer_alertas_eventos(panel.get('proximas', [])),
    }

# ==================================================
# HELPER ENVÍO ALERTAS EMAIL
# ==================================================

def _enviar_alerta_urgente_email_base():
    if not getattr(settings, 'ALERTAS_AUTOMATICAS_ACTIVAS', True):
        return {
            'ok': False,
            'enviado': False,
            'mensaje': 'Las alertas automáticas están desactivadas en settings.',
        }

    destinatarios = list(getattr(settings, 'ALERTAS_EMAIL_DESTINATARIOS', []))
    destinatarios = [d for d in destinatarios if d]

    if not destinatarios:
        return {
            'ok': False,
            'enviado': False,
            'mensaje': 'No hay destinatarios configurados en ALERTAS_EMAIL_DESTINATARIOS.',
        }

    dias_urgentes = getattr(settings, 'ALERTAS_URGENTES_DIAS', 2)
    limite_eventos = getattr(settings, 'ALERTAS_EMAIL_LIMITE_EVENTOS', 200)

    resumen = resumen_alertas_urgentes_email(dias=dias_urgentes)
    eventos = obtener_alertas_urgentes_email(dias=dias_urgentes, limite=limite_eventos)
    eventos = _enriquecer_alertas_eventos(eventos)

    if resumen['total_eventos'] == 0:
        return {
            'ok': True,
            'enviado': False,
            'mensaje': f'No existen alertas urgentes para enviar (vencidas + próximas en {dias_urgentes} día(s)).',
            'resumen': resumen,
            'destinatarios': destinatarios,
        }

    contexto = {
        'resumen': resumen,
        'eventos': eventos,
        'dias_urgentes': dias_urgentes,
        'fecha_generacion': timezone.localtime(),
    }

    html_content = render_to_string('pagos/emails/alerta_urgente_resumen.html', contexto)
    text_content = strip_tags(html_content)

    asunto_base = getattr(settings, 'ALERTAS_EMAIL_ASUNTO', 'Alerta financiera urgente')
    subject = f"{asunto_base}: {resumen['total_eventos']} compromiso(s) crítico(s) por saldo pendiente"

    email = EmailMultiAlternatives(
        subject=subject,
        body=text_content,
        from_email=getattr(settings, 'DEFAULT_FROM_EMAIL', None),
        to=destinatarios,
    )
    email.attach_alternative(html_content, "text/html")
    email.send(fail_silently=False)

    return {
        'ok': True,
        'enviado': True,
        'mensaje': f'Alerta urgente enviada correctamente a: {", ".join(destinatarios)}',
        'resumen': resumen,
        'destinatarios': destinatarios,
    }

# ==================================================
# HELPERS CONCILIACIÓN
# ==================================================

def _tokenize(texto):
    t = (texto or "").lower()
    t = re.sub(r"[^a-z0-9áéíóúñ ]", " ", t)
    parts = [p for p in t.split() if len(p) >= 3]
    return set(parts)


def _score_match(mov: MovimientoBancario, pago: PagoReal):
    score = 0

    if mov.monto == pago.monto:
        score += 60
    else:
        diff = abs(mov.monto - pago.monto)
        if diff <= Decimal("1.00"):
            score += 45
        elif diff <= Decimal("50.00"):
            score += 25

    delta = abs((mov.fecha - pago.fecha_pago).days)
    if delta == 0:
        score += 20
    elif delta <= 2:
        score += 14
    elif delta <= 5:
        score += 8

    a = _tokenize((mov.descripcion or "") + " " + (mov.referencia or ""))
    b = _tokenize((pago.pago.nombre or "") + " " + (pago.observacion or ""))
    if a and b:
        inter = len(a.intersection(b))
        if inter >= 3:
            score += 20
        elif inter == 2:
            score += 14
        elif inter == 1:
            score += 8

    return score


def _pago_ya_conciliado_con_otro_movimiento(pago_real_id, mov_id=None):
    qs = MovimientoBancario.objects.filter(
        conciliado=True,
        pago_real_id=pago_real_id
    )
    if mov_id:
        qs = qs.exclude(id=mov_id)
    return qs.exists()


def _obtener_candidatos_para_movimiento(mov, max_candidatos=25, excluir_pago_ids=None):
    excluir_pago_ids = excluir_pago_ids or set()

    if mov.tipo != "cargo":
        return []

    f1 = mov.fecha - timedelta(days=5)
    f2 = mov.fecha + timedelta(days=5)

    low = mov.monto - Decimal("50.00")
    high = mov.monto + Decimal("50.00")
    if low < 0:
        low = Decimal("0.00")

    candidatos = (
        PagoReal.objects
        .filter(fecha_pago__range=[f1, f2], monto__range=[low, high])
        .select_related("pago")
        .order_by("-fecha_pago", "-id")[:max_candidatos]
    )

    ranked = []
    for p in candidatos:
        if p.id in excluir_pago_ids:
            continue
        if _pago_ya_conciliado_con_otro_movimiento(p.id, mov_id=mov.id):
            continue

        s = _score_match(mov, p)
        if s >= 45:
            ranked.append((s, p))

    ranked.sort(key=lambda x: x[0], reverse=True)
    return ranked


# ==================================================
# HELPERS IMPORTACIÓN MASIVA EXCEL
# ==================================================

def _parse_fecha_excel_deuda(fecha_raw):
    if not fecha_raw:
        return None

    d = _parse_fecha_cartola(fecha_raw)
    if d:
        return d

    meses = {
        'enero': 1,
        'febrero': 2,
        'marzo': 3,
        'abril': 4,
        'mayo': 5,
        'junio': 6,
        'julio': 7,
        'agosto': 8,
        'septiembre': 9,
        'setiembre': 9,
        'octubre': 10,
        'noviembre': 11,
        'diciembre': 12,
    }

    s = str(fecha_raw).strip().lower()
    if s in meses:
        hoy = timezone.now().date()
        return date(hoy.year, meses[s], 1)

    return None


def _parse_cuotas_excel(raw):
    if raw is None:
        return None, None

    s = str(raw).strip()
    if not s:
        return None, None

    m = re.match(r'^\s*(\d+)\s*/\s*(\d+)\s*$', s)
    if not m:
        return None, None

    actual = int(m.group(1))
    total = int(m.group(2))

    if actual <= 0 or total <= 0 or actual > total:
        return None, None

    return actual, total


def _normalizar_tipo_programado_excel(tipo_excel, tipo_pago_excel):
    t1 = str(tipo_excel or "").strip().lower()
    t2 = str(tipo_pago_excel or "").strip().lower()
    combinado = f"{t1} {t2}"

    if "cheque" in combinado:
        return "cheque"
    if "credito" in combinado or "crédito" in combinado or "prestamo" in combinado or "préstamo" in combinado:
        return "credito"
    if "mensual" in combinado:
        return "fijo"
    return "unico"


def _normalizar_frecuencia_excel(tipo_pago_excel, total_cuotas):
    t = str(tipo_pago_excel or "").strip().lower()

    if "mensual" in t:
        return "mensual"
    if "quincenal" in t:
        return "quincenal"
    if "semanal" in t:
        return "semanal"
    if "unico" in t or "único" in t or "unica" in t or "única" in t:
        return "unico"

    if total_cuotas and total_cuotas > 1:
        return "mensual"

    return "unico"


def _metodo_pago_desde_texto(tipo_excel, tipo_pago_excel):
    t = f"{str(tipo_excel or '').lower()} {str(tipo_pago_excel or '').lower()}"

    if "cheque" in t:
        return "cheque"
    if "debito" in t or "débito" in t:
        return "debito"
    if "credito" in t or "crédito" in t or "tarjeta" in t:
        return "credito"
    if "efectivo" in t:
        return "efectivo"
    return "transferencia"


def _es_fila_resumen_importacion(id_deuda, descripcion, tipo_excel, monto):
    texto = f"{id_deuda or ''} {descripcion or ''} {tipo_excel or ''}".strip().lower()

    if not texto and (monto is None or str(monto).strip() == ""):
        return True

    if "total semana" in texto or "total vencidas" in texto or "deudas vencidas prioridad" in texto:
        return True

    return False


def _buscar_deuda_existente_importacion(nombre, fecha_inicio, monto, total_cuotas):
    return (
        PagoProgramado.objects
        .filter(
            nombre__iexact=(nombre or "").strip(),
            fecha_inicio=fecha_inicio,
            monto=monto,
            total_cuotas=total_cuotas,
        )
        .first()
    )


def _build_descripcion_importada(id_deuda, tipo_excel, tipo_pago_excel, cuotas_raw, saldo_raw):
    bloques = ["Importado desde Excel"]

    if id_deuda:
        bloques.append(f"ID origen: {id_deuda}")
    if tipo_excel:
        bloques.append(f"Tipo Excel: {tipo_excel}")
    if tipo_pago_excel:
        bloques.append(f"Tipo de pago Excel: {tipo_pago_excel}")
    if cuotas_raw:
        bloques.append(f"Cuotas Excel: {cuotas_raw}")
    if saldo_raw not in (None, ""):
        bloques.append(f"Saldo Excel al importar: {saldo_raw}")

    return "\n".join(bloques)


def _debt_preview_key(nombre, fecha_inicio, monto, total_cuotas):
    return f"{(nombre or '').strip().lower()}|{fecha_inicio.isoformat()}|{monto}|{total_cuotas}"


def _payment_preview_key(debt_key, fecha_pago, monto):
    return f"{debt_key}|{fecha_pago.isoformat()}|{monto}"


def _construir_preview_importacion(rows, crear_pagos_reales):
    resumen = {
        'deudas_nuevas': 0,
        'deudas_existentes': 0,
        'pagos_nuevos': 0,
        'pagos_existentes': 0,
        'omitidas': 0,
        'errores': 0,
        'fechas_estimadas': 0,
        'filas_procesadas': 0,
    }

    preview_rows = []
    entries_confirmacion = []

    debt_keys_nuevas = set()
    payment_keys_nuevos = set()

    for idx, r in enumerate(rows, start=2):
        try:
            resumen['filas_procesadas'] += 1

            id_deuda = str(r.get('id_deuda') or '').strip()
            fecha_raw = r.get('FECHA DE PAGO')
            nombre = str(r.get('DESCRIPCION') or '').strip()
            tipo_excel = str(r.get('TIPO') or '').strip()
            monto_raw = r.get('MONTO')
            pagado_raw = r.get('PAGADO')
            saldo_raw = r.get('SALDOS')
            cuotas_raw = r.get('CUOTAS')
            tipo_pago_excel = str(r.get('tipo de pago') or '').strip()

            if _es_fila_resumen_importacion(id_deuda, nombre, tipo_excel, monto_raw):
                resumen['omitidas'] += 1
                if len(preview_rows) < 50:
                    preview_rows.append({
                        'fila': idx,
                        'descripcion': nombre or '—',
                        'fecha_inicio': '—',
                        'monto': '0',
                        'pagado': '0',
                        'cuotas_texto': cuotas_raw or '—',
                        'tipo': tipo_excel or '—',
                        'frecuencia': '—',
                        'estado_deuda': 'Omitida',
                        'estado_pago': '—',
                        'motivo': 'Fila de resumen / subtotal / vacía',
                    })
                continue

            if not nombre:
                resumen['omitidas'] += 1
                if len(preview_rows) < 50:
                    preview_rows.append({
                        'fila': idx,
                        'descripcion': '—',
                        'fecha_inicio': '—',
                        'monto': '0',
                        'pagado': '0',
                        'cuotas_texto': cuotas_raw or '—',
                        'tipo': tipo_excel or '—',
                        'frecuencia': '—',
                        'estado_deuda': 'Omitida',
                        'estado_pago': '—',
                        'motivo': 'Sin descripción',
                    })
                continue

            monto = _to_decimal_monto(monto_raw)
            pagado = _to_decimal_monto(pagado_raw)
            saldo = _to_decimal_monto(saldo_raw)

            if monto <= 0:
                resumen['omitidas'] += 1
                if len(preview_rows) < 50:
                    preview_rows.append({
                        'fila': idx,
                        'descripcion': nombre,
                        'fecha_inicio': '—',
                        'monto': str(monto),
                        'pagado': str(pagado),
                        'cuotas_texto': cuotas_raw or '—',
                        'tipo': tipo_excel or '—',
                        'frecuencia': '—',
                        'estado_deuda': 'Omitida',
                        'estado_pago': '—',
                        'motivo': 'Monto principal inválido o cero',
                    })
                continue

            fecha_inicio = _parse_fecha_excel_deuda(fecha_raw)
            fecha_estimada = False
            if not fecha_inicio:
                fecha_inicio = timezone.now().date()
                fecha_estimada = True
                resumen['fechas_estimadas'] += 1

            cuota_actual, total_cuotas_excel = _parse_cuotas_excel(cuotas_raw)

            if total_cuotas_excel:
                total_cuotas = total_cuotas_excel
                cuotas_restantes = max(total_cuotas_excel - cuota_actual, 0)
            else:
                total_cuotas = 1
                cuotas_restantes = 1

            tipo_programado = _normalizar_tipo_programado_excel(tipo_excel, tipo_pago_excel)
            frecuencia = _normalizar_frecuencia_excel(tipo_pago_excel, total_cuotas)
            debt_key = _debt_preview_key(nombre, fecha_inicio, monto, total_cuotas)

            deuda_existente = _buscar_deuda_existente_importacion(
                nombre=nombre,
                fecha_inicio=fecha_inicio,
                monto=monto,
                total_cuotas=total_cuotas,
            )

            if deuda_existente:
                estado_deuda = 'Existente BD'
                deuda_existente_id = deuda_existente.id
                resumen['deudas_existentes'] += 1
            else:
                deuda_existente_id = None
                if debt_key in debt_keys_nuevas:
                    estado_deuda = 'Repetida en archivo'
                else:
                    debt_keys_nuevas.add(debt_key)
                    estado_deuda = 'Nueva'
                    resumen['deudas_nuevas'] += 1

            estado_pago = 'Sin pago'
            crear_pago = False
            metodo_pago = _metodo_pago_desde_texto(tipo_excel, tipo_pago_excel)
            payment_key = None

            if crear_pagos_reales and pagado > 0:
                payment_key = _payment_preview_key(debt_key, fecha_inicio, pagado)

                pago_existente = None
                if deuda_existente:
                    pago_existente = PagoReal.objects.filter(
                        pago=deuda_existente,
                        fecha_pago=fecha_inicio,
                        monto=pagado,
                    ).first()

                if pago_existente:
                    estado_pago = 'Existente BD'
                    resumen['pagos_existentes'] += 1
                else:
                    if payment_key in payment_keys_nuevos:
                        estado_pago = 'Repetido en archivo'
                    else:
                        payment_keys_nuevos.add(payment_key)
                        estado_pago = 'Nuevo'
                        crear_pago = True
                        resumen['pagos_nuevos'] += 1

            descripcion_importada = _build_descripcion_importada(
                id_deuda=id_deuda,
                tipo_excel=tipo_excel,
                tipo_pago_excel=tipo_pago_excel,
                cuotas_raw=cuotas_raw,
                saldo_raw=saldo_raw,
            )

            entry = {
                'fila': idx,
                'id_deuda': id_deuda,
                'nombre': nombre[:120],
                'fecha_inicio': fecha_inicio.isoformat(),
                'monto': str(monto),
                'pagado': str(pagado),
                'saldo': str(saldo),
                'cuotas_raw': str(cuotas_raw or ''),
                'tipo_excel': tipo_excel,
                'tipo_pago_excel': tipo_pago_excel,
                'tipo_programado': tipo_programado,
                'frecuencia': frecuencia,
                'total_cuotas': total_cuotas,
                'cuotas_restantes': cuotas_restantes,
                'descripcion_importada': descripcion_importada,
                'debt_key': debt_key,
                'deuda_existente_id': deuda_existente_id,
                'crear_pago': crear_pago,
                'metodo_pago': metodo_pago,
                'payment_key': payment_key,
                'fecha_estimada': fecha_estimada,
                'estado_deuda': estado_deuda,
                'estado_pago': estado_pago,
            }

            entries_confirmacion.append(entry)

            if len(preview_rows) < 50:
                preview_rows.append({
                    'fila': idx,
                    'descripcion': nombre,
                    'fecha_inicio': fecha_inicio.strftime('%d-%m-%Y'),
                    'monto': f"{monto:,.0f}",
                    'pagado': f"{pagado:,.0f}",
                    'cuotas_texto': cuotas_raw or "1/1",
                    'tipo': tipo_programado,
                    'frecuencia': frecuencia,
                    'estado_deuda': estado_deuda,
                    'estado_pago': estado_pago,
                    'motivo': 'Fecha estimada automáticamente' if fecha_estimada else '—',
                })

        except Exception as e:
            resumen['errores'] += 1
            if len(preview_rows) < 50:
                preview_rows.append({
                    'fila': idx,
                    'descripcion': str(r.get('DESCRIPCION') or '—'),
                    'fecha_inicio': '—',
                    'monto': '—',
                    'pagado': '—',
                    'cuotas_texto': str(r.get('CUOTAS') or '—'),
                    'tipo': str(r.get('TIPO') or '—'),
                    'frecuencia': '—',
                    'estado_deuda': 'Error',
                    'estado_pago': '—',
                    'motivo': str(e),
                })

    return {
        'resumen': resumen,
        'preview_rows': preview_rows,
        'entries_confirmacion': entries_confirmacion,
    }


# ==================================================
# DASHBOARD FINANCIERO EJECUTIVO
# ==================================================

@staff_member_required
def dashboard_financiero(request):
    kpis = obtener_kpis_financieros()

    flujo_chart = flujo_proyectado_mensual_chart(6)
    flujo_chart_json = json.dumps(flujo_chart)

    riesgo_data = calcular_riesgo_financiero()
    riesgo = riesgo_data.get('nivel', 'N/A')
    porcentaje = riesgo_data.get('porcentaje', 0)

    vencidos = eventos_vencidos()
    proximos = eventos_proximos(dias=7)

    alertas_resumen = resumen_alertas_financieras()
    pagos = listar_compromisos_financieros(include_pagados=True)
    resumen_estados = resumen_estados_compromisos()

    contexto = {
        'kpis': kpis,
        'flujo_chart_json': flujo_chart_json,
        'riesgo': riesgo,
        'porcentaje': porcentaje,
        'eventos_vencidos': vencidos,
        'eventos_proximos': proximos,
        'alertas_resumen': alertas_resumen,
        'pagos': pagos,
        'resumen_estados': resumen_estados,
    }

    return render(request, 'pagos/dashboard.html', contexto)


@staff_member_required
def alertas_financieras(request):
    dias_urgentes = getattr(settings, 'ALERTAS_URGENTES_DIAS', 2)
    panel_base = obtener_panel_alertas_financieras(limite=50)
    panel_enriquecido = _enriquecer_panel_alertas(panel_base)

    contexto = {
        'alertas_resumen': resumen_alertas_financieras(),
        'alertas_panel': panel_enriquecido,
        'alertas_email_resumen': resumen_alertas_urgentes_email(dias=dias_urgentes),
        'dias_urgentes': dias_urgentes,
        'puede_enviar_alerta_email': (
            request.user.is_superuser or request.user.has_perm('pagos.change_pagoprogramado')
        ),
    }
    return render(request, 'pagos/alertas_financieras.html', contexto)


@staff_member_required
def enviar_alerta_urgente_email(request):
    if request.method != 'POST':
        return redirect('alertas_financieras')

    if not (request.user.is_superuser or request.user.has_perm('pagos.change_pagoprogramado')):
        messages.error(request, 'No tienes permisos para enviar alertas por correo.')
        return redirect('alertas_financieras')

    try:
        resultado = _enviar_alerta_urgente_email_base()

        if resultado['ok'] and resultado['enviado']:
            messages.success(request, resultado['mensaje'])
        elif resultado['ok'] and not resultado['enviado']:
            messages.info(request, resultado['mensaje'])
        else:
            messages.warning(request, resultado['mensaje'])

    except Exception as e:
        messages.error(request, f'No se pudo enviar el correo de alertas: {e}')

    return redirect('alertas_financieras')


# ==================================================
# LISTADO DE CUENTAS POR PAGAR
# ==================================================

@staff_member_required
def pagos_lista(request):
    q = (request.GET.get('q') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    estado = (request.GET.get('estado') or '').strip().upper()
    activo_raw = (request.GET.get('activo') or '').strip().lower()
    ver_pagados = (request.GET.get('ver_pagados') or '').strip() == '1'

    activo = None
    if activo_raw == '1':
        activo = True
    elif activo_raw == '0':
        activo = False

    pagos_base = listar_compromisos_financieros(
        include_pagados=ver_pagados,
        q=q,
        tipo=tipo,
        estado=estado,
        activo=activo,
    )

    pagos = []
    for item in pagos_base:
        pagos.append({
            'id': item['id'],
            'fecha_inicio': item['fecha_inicio'],
            'nombre': item['nombre'],
            'tipo': item['tipo'],
            'total': item['total_compromiso'],
            'pagado': item['pagado_real'],
            'saldo': item['saldo_real'],
            'estado': item['estado_real'],
            'porcentaje': item['porcentaje_pagado'],
            'activo': item['activo'],
            'total_cuotas': item['total_cuotas'],
            'cuotas_restantes': item['cuotas_restantes'],
        })

    resumen_operativo = {
        'total': len(pagos),
        'pendientes': sum(1 for p in pagos if p['estado'] == 'PENDIENTE'),
        'parciales': sum(1 for p in pagos if p['estado'] == 'PARCIAL'),
        'pagados': sum(1 for p in pagos if p['estado'] == 'PAGADO'),
    }

    tipos_disponibles = [
        {'value': value, 'label': label}
        for value, label in PagoProgramado.TIPO_CHOICES
    ]

    contexto = {
        'pagos': pagos,
        'tipos_disponibles': tipos_disponibles,
        'filtros': {
            'q': q,
            'tipo': tipo,
            'estado': estado,
            'activo': activo_raw,
            'ver_pagados': ver_pagados,
        },
        'resumen_operativo': resumen_operativo,
    }

    return render(request, 'pagos/pagos_lista.html', contexto)


# ==================================================
# CREAR / EDITAR PAGO PROGRAMADO
# ==================================================

@staff_member_required
def pagos_crear(request):
    if request.method == 'POST':
        form = PagoProgramadoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago registrado correctamente.')
            return redirect('pagos_lista')
        messages.error(request, 'Revisa el formulario, hay campos inválidos.')
    else:
        form = PagoProgramadoForm()

    return render(request, 'pagos/pagos_form.html', {
        'form': form,
        'titulo': 'Nuevo compromiso / deuda',
        'modo_edicion': False,
    })


@staff_member_required
def pagos_editar(request, pk):
    pago = get_object_or_404(PagoProgramado, pk=pk)

    if request.method == 'POST':
        form = PagoProgramadoForm(request.POST, instance=pago)
        if form.is_valid():
            form.save()
            messages.success(request, 'Deuda actualizada correctamente.')

            if pago.pagos_realizados.exists() or pago.eventos.exists():
                messages.warning(
                    request,
                    'Ojo: esta deuda ya tiene pagos o eventos asociados. '
                    'Si cambiaste monto, cuotas o frecuencia, revisa saldos y proyecciones.'
                )

            return redirect('pagos_lista')

        messages.error(request, 'Revisa el formulario, hay campos inválidos.')
    else:
        form = PagoProgramadoForm(instance=pago)

        if pago.pagos_realizados.exists() or pago.eventos.exists():
            messages.info(
                request,
                'Esta deuda ya tiene pagos o eventos asociados. '
                'Edita con cuidado monto, cuotas, fecha y frecuencia.'
            )

    return render(request, 'pagos/pagos_form.html', {
        'form': form,
        'titulo': f'Editar deuda: {pago.nombre}',
        'modo_edicion': True,
        'pago_obj': pago,
    })


# ==================================================
# IMPORTACIÓN MASIVA EXCEL
# ==================================================

@staff_member_required
def pagos_importar_excel(request):
    preview_data = request.session.get('pagos_import_preview')

    if request.method == 'POST':
        form = PagosImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']
            hoja = (form.cleaned_data.get('hoja') or '').strip()
            crear_pagos_reales = form.cleaned_data.get('crear_pagos_reales')

            name = (archivo.name or '').lower()

            try:
                archivo.seek(0)
                if name.endswith('.xlsx'):
                    rows = _parse_xlsx(archivo, sheet_name=hoja or None)
                else:
                    rows = _parse_csv(archivo, sep=';')
            except Exception as e:
                messages.error(request, f'No se pudo leer el archivo de importación: {e}')
                return render(request, 'pagos/pagos_importar_excel.html', {
                    'form': form,
                    'preview_data': preview_data,
                })

            resultado_preview = _construir_preview_importacion(rows, crear_pagos_reales)

            preview_data = {
                'archivo_nombre': archivo.name,
                'hoja': hoja,
                'crear_pagos_reales': crear_pagos_reales,
                'resumen': resultado_preview['resumen'],
                'preview_rows': resultado_preview['preview_rows'],
                'entries_confirmacion': resultado_preview['entries_confirmacion'],
            }

            request.session['pagos_import_preview'] = preview_data
            request.session.modified = True

            messages.info(
                request,
                'Vista previa generada correctamente. Revisa los datos antes de confirmar la importación.'
            )

            return render(request, 'pagos/pagos_importar_excel.html', {
                'form': PagosImportExcelForm(initial={
                    'hoja': hoja,
                    'crear_pagos_reales': crear_pagos_reales,
                }),
                'preview_data': preview_data,
            })

        messages.error(request, 'Revisa el formulario de importación.')
    else:
        form = PagosImportExcelForm()

    return render(request, 'pagos/pagos_importar_excel.html', {
        'form': form,
        'preview_data': preview_data,
    })


@staff_member_required
def pagos_importar_excel_limpiar_preview(request):
    if 'pagos_import_preview' in request.session:
        del request.session['pagos_import_preview']
        request.session.modified = True
        messages.info(request, 'Vista previa eliminada.')

    destino = request.GET.get('next') or 'pagos_importar_excel'
    return redirect(destino)


@staff_member_required
def pagos_importar_excel_confirmar(request):
    if request.method != 'POST':
        return redirect('pagos_importar_excel')

    preview_data = request.session.get('pagos_import_preview')
    if not preview_data:
        messages.warning(request, 'No hay una vista previa disponible para confirmar.')
        return redirect('pagos_importar_excel')

    entries = preview_data.get('entries_confirmacion', [])

    creadas = 0
    existentes = 0
    pagos_creados = 0
    pagos_existentes = 0
    omitidas = preview_data.get('resumen', {}).get('omitidas', 0)
    errores = 0
    fechas_estimadas = preview_data.get('resumen', {}).get('fechas_estimadas', 0)

    cache_deudas_archivo = {}
    cache_pagos_archivo = set()

    importacion = ImportacionPago.objects.create(
        usuario=request.user if request.user.is_authenticated else None,
        archivo_nombre=preview_data.get('archivo_nombre', 'archivo_importado'),
        hoja=preview_data.get('hoja', '') or '',
        crear_pagos_reales=preview_data.get('crear_pagos_reales', False),
        resumen=preview_data.get('resumen', {}),
        total_deudas_creadas=0,
        total_deudas_existentes=0,
        total_pagos_creados=0,
        total_pagos_existentes=0,
        total_omitidas=omitidas,
        total_errores=0,
        estado='confirmada',
    )

    for entry in entries:
        try:
            nombre = entry['nombre']
            fecha_inicio = parse_date(entry['fecha_inicio'])
            monto = Decimal(entry['monto'])
            pagado = Decimal(entry['pagado'])
            saldo = Decimal(entry['saldo'])
            total_cuotas = int(entry['total_cuotas'])
            cuotas_restantes = int(entry['cuotas_restantes'])
            tipo_programado = entry['tipo_programado']
            frecuencia = entry['frecuencia']
            descripcion_importada = entry['descripcion_importada']
            tipo_excel = entry['tipo_excel']
            tipo_pago_excel = entry['tipo_pago_excel']
            id_deuda = entry['id_deuda']
            debt_key = entry['debt_key']
            metodo_pago = entry['metodo_pago']
            payment_key = entry.get('payment_key')
            deuda_existente_id = entry.get('deuda_existente_id')
            crear_pago = entry.get('crear_pago', False)
            fila_excel = entry.get('fila')

            deuda = None

            if deuda_existente_id:
                deuda = PagoProgramado.objects.filter(pk=deuda_existente_id).first()
                if deuda:
                    existentes += 1

            if deuda is None and debt_key in cache_deudas_archivo:
                deuda = cache_deudas_archivo[debt_key]

            if deuda is None:
                deuda = _buscar_deuda_existente_importacion(
                    nombre=nombre,
                    fecha_inicio=fecha_inicio,
                    monto=monto,
                    total_cuotas=total_cuotas,
                )

                if deuda:
                    existentes += 1
                else:
                    deuda = PagoProgramado.objects.create(
                        nombre=nombre[:120],
                        tipo=tipo_programado,
                        monto=monto,
                        fecha_inicio=fecha_inicio,
                        frecuencia=frecuencia,
                        total_cuotas=total_cuotas,
                        cuotas_restantes=cuotas_restantes,
                        descripcion=descripcion_importada,
                        activo=True,
                    )
                    creadas += 1

                    ImportacionPagoDetalle.objects.create(
                        importacion=importacion,
                        fila_excel=fila_excel,
                        tipo_registro='deuda',
                        pago_programado=deuda,
                        descripcion=nombre[:180],
                    )

            cache_deudas_archivo[debt_key] = deuda

            if crear_pago and pagado > 0:
                if payment_key and payment_key in cache_pagos_archivo:
                    continue

                pago_existente = PagoReal.objects.filter(
                    pago=deuda,
                    fecha_pago=fecha_inicio,
                    monto=pagado,
                ).first()

                if pago_existente:
                    pagos_existentes += 1
                else:
                    pago_real = PagoReal.objects.create(
                        pago=deuda,
                        fecha_pago=fecha_inicio,
                        monto=pagado,
                        metodo_pago=metodo_pago,
                        observacion=(
                            f"Importado desde Excel"
                            f"{f' | ID origen: {id_deuda}' if id_deuda else ''}"
                            f"{f' | Saldo Excel: {saldo}' if saldo else ''}"
                            f"{f' | Tipo Excel: {tipo_excel}' if tipo_excel else ''}"
                            f"{f' | Tipo de pago Excel: {tipo_pago_excel}' if tipo_pago_excel else ''}"
                        )
                    )
                    pagos_creados += 1

                    ImportacionPagoDetalle.objects.create(
                        importacion=importacion,
                        fila_excel=fila_excel,
                        tipo_registro='pago_real',
                        pago_programado=deuda,
                        pago_real=pago_real,
                        descripcion=nombre[:180],
                    )

                if payment_key:
                    cache_pagos_archivo.add(payment_key)

        except Exception:
            errores += 1
            continue

    importacion.total_deudas_creadas = creadas
    importacion.total_deudas_existentes = existentes
    importacion.total_pagos_creados = pagos_creados
    importacion.total_pagos_existentes = pagos_existentes
    importacion.total_omitidas = omitidas
    importacion.total_errores = errores
    importacion.resumen = {
        **(preview_data.get('resumen', {}) or {}),
        'deudas_nuevas_confirmadas': creadas,
        'deudas_existentes_confirmadas': existentes,
        'pagos_nuevos_confirmados': pagos_creados,
        'pagos_existentes_confirmados': pagos_existentes,
        'omitidas_confirmadas': omitidas,
        'errores_confirmados': errores,
    }
    importacion.save()

    if 'pagos_import_preview' in request.session:
        del request.session['pagos_import_preview']
        request.session.modified = True

    messages.success(
        request,
        f'Importación de deudas confirmada ✅ '
        f'Deudas nuevas: {creadas} | '
        f'Deudas existentes: {existentes} | '
        f'Pagos creados: {pagos_creados} | '
        f'Pagos existentes: {pagos_existentes} | '
        f'Omitidas: {omitidas} | '
        f'Errores: {errores}'
    )

    messages.info(
        request,
        f'La importación quedó registrada en el historial con ID #{importacion.id}.'
    )

    if fechas_estimadas > 0:
        messages.warning(
            request,
            f'Se asignó fecha automática a {fechas_estimadas} fila(s) sin fecha reconocible. '
            f'Revisa esos registros luego de importar.'
        )

    return redirect('pagos_lista')


@staff_member_required
def importaciones_historial(request):
    importaciones = (
        ImportacionPago.objects
        .select_related('usuario', 'revertida_por')
        .prefetch_related('detalles')
        .order_by('-creado')
    )

    contexto = {
        'importaciones': importaciones,
        'kpi_total': importaciones.count(),
        'kpi_confirmadas': importaciones.filter(estado='confirmada').count(),
        'kpi_revertidas': importaciones.filter(estado='revertida').count(),
    }
    return render(request, 'pagos/importaciones_historial.html', contexto)


@staff_member_required
def importacion_revertir(request, pk):
    if request.method != 'POST':
        return redirect('importaciones_historial')

    importacion = get_object_or_404(
        ImportacionPago.objects.prefetch_related('detalles'),
        pk=pk
    )

    if importacion.estado == 'revertida':
        messages.info(request, 'Esta importación ya fue revertida anteriormente.')
        return redirect('importaciones_historial')

    detalles = list(importacion.detalles.all())

    pago_ids = sorted({d.pago_real_id for d in detalles if d.tipo_registro == 'pago_real' and d.pago_real_id})
    deuda_ids = sorted({d.pago_programado_id for d in detalles if d.tipo_registro == 'deuda' and d.pago_programado_id})

    pagos_eliminados = 0
    deudas_eliminadas = 0

    with transaction.atomic():
        if pago_ids:
            pagos_qs = PagoReal.objects.filter(id__in=pago_ids)
            pagos_eliminados = pagos_qs.count()
            pagos_qs.delete()

        if deuda_ids:
            deudas_qs = PagoProgramado.objects.filter(id__in=deuda_ids)
            deudas_eliminadas = deudas_qs.count()
            deudas_qs.delete()

        importacion.estado = 'revertida'
        importacion.revertida_en = timezone.now()
        importacion.revertida_por = request.user if request.user.is_authenticated else None
        importacion.save(update_fields=['estado', 'revertida_en', 'revertida_por'])

    messages.success(
        request,
        f'Importación #{importacion.id} revertida correctamente ✅ '
        f'Deudas eliminadas: {deudas_eliminadas} | '
        f'Pagos eliminados: {pagos_eliminados}'
    )
    return redirect('importaciones_historial')


# ==================================================
# REGISTRAR / EDITAR PAGO REAL
# ==================================================

@staff_member_required
def pagos_real_crear(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'dashboard_financiero'
    pago_id = request.GET.get('pago_id')

    if request.method == 'POST':
        form = PagoRealForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago registrado correctamente.')
            return redirect(next_url)
        messages.error(request, 'Revisa el formulario, hay campos inválidos.')
    else:
        initial = {}
        if pago_id:
            initial['pago'] = pago_id
        form = PagoRealForm(initial=initial)

    return render(request, 'pagos/pagos_real_form.html', {
        'form': form,
        'titulo': 'Registrar pago real',
        'next': next_url,
        'modo_edicion': False,
        'pago_real_obj': None,
        'tiene_conciliacion': False,
        'movimiento_conciliado': None,
    })


@staff_member_required
def pagos_real_editar(request, pk):
    pago_real = get_object_or_404(
        PagoReal.objects.select_related('pago'),
        pk=pk
    )

    next_url = request.GET.get('next') or request.POST.get('next') or 'reportes_financieros'

    movimiento_conciliado = (
        pago_real.movimientos_conciliados
        .filter(conciliado=True)
        .order_by('-conciliado_en', '-id')
        .first()
    )

    tiene_conciliacion = movimiento_conciliado is not None

    if request.method == 'POST':
        form = PagoRealForm(request.POST, instance=pago_real)
        if form.is_valid():
            form.save()
            messages.success(request, 'Pago real actualizado correctamente.')

            if tiene_conciliacion:
                messages.warning(
                    request,
                    'Se modificó un pago con conciliación existente. '
                    'Revisa la conciliación bancaria para confirmar que siga consistente.'
                )

            return redirect(next_url)

        messages.error(request, 'Revisa el formulario, hay campos inválidos.')
    else:
        form = PagoRealForm(instance=pago_real)

        if tiene_conciliacion:
            messages.info(
                request,
                'Este pago está conciliado con un movimiento bancario. '
                'Edita con cuidado monto, fecha, compromiso y observación.'
            )

    return render(request, 'pagos/pagos_real_form.html', {
        'form': form,
        'titulo': f'Editar pago real: {pago_real.pago.nombre}',
        'next': next_url,
        'modo_edicion': True,
        'pago_real_obj': pago_real,
        'tiene_conciliacion': tiene_conciliacion,
        'movimiento_conciliado': movimiento_conciliado,
    })


# ==================================================
# REPORTES FINANCIEROS (V2) + EXPORT
# ==================================================

@staff_member_required
def reportes_financieros(request):
    desde, hasta = _get_rango_fechas_from_request(request)
    pagos_qs = _build_report_queryset(desde, hasta)

    total = pagos_qs.aggregate(
        total=Coalesce(
            Sum('monto'),
            Value(Decimal('0.00')),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    )['total']

    count = pagos_qs.count()
    promedio = (total / Decimal(count)) if count else Decimal('0.00')

    export = (request.GET.get('export') or "").lower().strip()
    if request.method == "GET" and export in ("csv", "xlsx", "pdf"):
        if export == "csv":
            return _export_csv(pagos_qs, desde, hasta)
        if export == "xlsx":
            return _export_xlsx(pagos_qs, desde, hasta)
        return _export_pdf(pagos_qs, desde, hasta, total, promedio)

    chart_diario_json = "{}"
    chart_metodo_json = "{}"
    top_compromisos = []
    metodo_principal = "—"

    if request.method == 'POST':
        diarios = (
            pagos_qs
            .values('fecha_pago')
            .annotate(
                total_dia=Coalesce(
                    Sum('monto'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=14, decimal_places=2)
                )
            )
            .order_by('fecha_pago')
        )

        chart_diario_json = json.dumps({
            "labels": [d['fecha_pago'].strftime('%Y-%m-%d') for d in diarios if d.get('fecha_pago')],
            "valores": [float(d['total_dia']) for d in diarios if d.get('fecha_pago')],
        })

        por_metodo = list(
            pagos_qs
            .values('metodo_pago')
            .annotate(
                total_metodo=Coalesce(
                    Sum('monto'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=14, decimal_places=2)
                )
            )
            .order_by('-total_metodo')
        )

        chart_metodo_json = json.dumps({
            "labels": [(m.get('metodo_pago') or "—") for m in por_metodo],
            "valores": [float(m['total_metodo']) for m in por_metodo],
        })

        if por_metodo:
            metodo_principal = (por_metodo[0].get('metodo_pago') or "—")

        top = (
            pagos_qs
            .values('pago__nombre')
            .annotate(
                total_comp=Coalesce(
                    Sum('monto'),
                    Value(Decimal('0.00')),
                    output_field=DecimalField(max_digits=14, decimal_places=2)
                ),
                cantidad=Count('id')
            )
            .order_by('-total_comp')[:3]
        )

        top_compromisos = [{
            "nombre": t['pago__nombre'],
            "total": t['total_comp'],
            "cantidad": t['cantidad'],
        } for t in top]

        form = ReportesFiltroForm(initial={"fecha_desde": desde, "fecha_hasta": hasta})
    else:
        form = ReportesFiltroForm(initial={"fecha_desde": desde, "fecha_hasta": hasta})

    return render(request, 'pagos/reportes.html', {
        'pagos': pagos_qs,
        'eventos': pagos_qs,
        'total': total,
        'promedio': promedio,
        'desde': desde,
        'hasta': hasta,
        'form': form,
        'chart_diario_json': chart_diario_json,
        'chart_metodo_json': chart_metodo_json,
        'top_compromisos': top_compromisos,
        'metodo_principal': metodo_principal,
    })


# ==================================================
# CARTOLAS - PARSEO (CSV/XLSX) + DEDUPE
# ==================================================

def _to_decimal_monto(raw):
    if raw is None or raw == "":
        return Decimal("0.00")

    if isinstance(raw, (int, float, Decimal)):
        try:
            return Decimal(str(raw))
        except Exception:
            return Decimal("0.00")

    s = str(raw).strip()

    s = s.replace("$", "").replace("CLP", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(",", "")
    else:
        if "," in s and "." not in s:
            s = s.replace(".", "")
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")

    try:
        return Decimal(s)
    except Exception:
        return Decimal("0.00")


def _normalizar_tipo(raw_tipo, monto):
    t = (str(raw_tipo or "").strip().lower())
    if t in ("cargo", "c", "debe", "debito", "débito", "egreso", "salida"):
        return "cargo"
    if t in ("abono", "a", "haber", "credito", "crédito", "ingreso", "entrada"):
        return "abono"

    if monto < 0:
        return "cargo"
    if monto > 0:
        return "abono"
    return "desconocido"


def _parse_fecha_cartola(fecha_raw):
    if not fecha_raw:
        return None

    if isinstance(fecha_raw, datetime):
        return fecha_raw.date()

    if isinstance(fecha_raw, date):
        return fecha_raw

    s = str(fecha_raw).strip()
    if not s:
        return None

    d = parse_date(s)
    if d:
        return d

    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d.%m.%Y", "%Y/%m/%d", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue

    return None


def _parse_csv(file_obj, sep=";"):
    raw = file_obj.read()
    if isinstance(raw, bytes):
        try:
            text = raw.decode("utf-8-sig")
        except Exception:
            text = raw.decode("latin-1")
    else:
        text = str(raw)

    f = io.StringIO(text)
    reader = csv.DictReader(f, delimiter=sep)
    return list(reader)


def _parse_xlsx(file_obj, sheet_name=None):
    from openpyxl import load_workbook

    wb = load_workbook(file_obj, data_only=True)

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.active

    headers = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(h).strip() if h is not None else "" for h in row]
            continue
        data = {}
        for idx, val in enumerate(row):
            key = headers[idx] if idx < len(headers) else f"col{idx}"
            data[key] = val
        rows.append(data)

    return rows


def _norm_key(k: str) -> str:
    return str(k or "").strip().lower().replace(" ", "")


def _find_key(row, wanted):
    if not wanted:
        return None

    w = _norm_key(wanted)

    for k in row.keys():
        if _norm_key(k) == w:
            return k

    for k in row.keys():
        kk = _norm_key(k)
        if w and w in kk:
            return k

    return None


@staff_member_required
def cartolas_importar(request):
    if request.method == "POST":
        form = CartolaImportForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            cuenta = (form.cleaned_data.get("cuenta") or "").strip()
            banco = (form.cleaned_data.get("banco") or "").strip()

            col_fecha = (form.cleaned_data.get("col_fecha") or "fecha").strip()
            col_descripcion = (form.cleaned_data.get("col_descripcion") or "descripcion").strip()
            col_monto = (form.cleaned_data.get("col_monto") or "monto").strip()
            col_referencia = (form.cleaned_data.get("col_referencia") or "referencia").strip()
            col_tipo = (form.cleaned_data.get("col_tipo") or "tipo").strip()
            sep = form.cleaned_data.get("separador_csv") or ";"

            name = (archivo.name or "").lower()
            try:
                archivo.seek(0)
                if name.endswith(".xlsx"):
                    rows = _parse_xlsx(archivo)
                else:
                    rows = _parse_csv(archivo, sep=sep)
            except Exception as e:
                messages.error(request, f"No se pudo leer el archivo: {e}")
                return render(request, "pagos/cartolas_importar.html", {"form": form})

            creados = 0
            duplicados = 0
            errores = 0

            for r in rows:
                try:
                    kf = _find_key(r, col_fecha)
                    kd = _find_key(r, col_descripcion)
                    km = _find_key(r, col_monto)
                    kr = _find_key(r, col_referencia) if col_referencia else None
                    kt = _find_key(r, col_tipo) if col_tipo else None

                    if not kf:
                        kf = _find_key(r, "Fecha")
                    if not kd:
                        kd = _find_key(r, "Detalle Movimiento")

                    fecha_raw = r.get(kf) if kf else None
                    desc_raw = r.get(kd) if kd else ""
                    monto_raw = r.get(km) if km else None
                    ref_raw = r.get(kr) if kr else ""
                    tipo_raw = r.get(kt) if kt else ""

                    fecha = _parse_fecha_cartola(fecha_raw)
                    if not fecha:
                        errores += 1
                        continue

                    descripcion = str(desc_raw or "").strip() or "—"

                    if km:
                        monto = _to_decimal_monto(monto_raw)
                    else:
                        k_cargo = _find_key(r, "Cheque o Cargo")
                        k_abono = _find_key(r, "Deposito o Abono")

                        cargo = _to_decimal_monto(r.get(k_cargo)) if k_cargo else Decimal("0.00")
                        abono = _to_decimal_monto(r.get(k_abono)) if k_abono else Decimal("0.00")

                        monto = abono - cargo

                    tipo = _normalizar_tipo(tipo_raw, monto)
                    monto_abs = abs(monto) if monto < 0 else monto

                    hu = MovimientoBancario.build_hash(
                        fecha=fecha.isoformat(),
                        monto=str(monto_abs),
                        descripcion=descripcion,
                        referencia=str(ref_raw or ""),
                        cuenta=cuenta
                    )

                    obj, created = MovimientoBancario.objects.get_or_create(
                        hash_unico=hu,
                        defaults={
                            "cuenta": cuenta,
                            "banco": banco,
                            "fecha": fecha,
                            "descripcion": descripcion[:255],
                            "referencia": str(ref_raw or "")[:120],
                            "tipo": tipo,
                            "monto": monto_abs,
                            "moneda": "CLP",
                            "raw": r,
                            "conciliado": False,
                            "pago_real": None,
                            "conciliado_en": None,
                            "nota_conciliacion": "",
                        }
                    )

                    if created:
                        creados += 1
                    else:
                        duplicados += 1

                except Exception:
                    errores += 1
                    continue

            if creados == 0 and duplicados == 0:
                messages.warning(
                    request,
                    f"No se importó ningún movimiento. Revisa mapeo de columnas. Filas con error: {errores}"
                )
                return render(request, "pagos/cartolas_importar.html", {"form": form})

            messages.success(
                request,
                f"Importación lista ✅ Nuevos: {creados} | Duplicados: {duplicados} | Filas con error: {errores}"
            )
            return redirect("cartolas_lista")

        messages.error(request, "Revisa el formulario de importación.")
    else:
        form = CartolaImportForm()

    return render(request, "pagos/cartolas_importar.html", {"form": form})


@staff_member_required
def cartolas_lista(request):
    qs = MovimientoBancario.objects.select_related("pago_real", "pago_real__pago").all()

    q = (request.GET.get("q") or "").strip()
    if q:
        qs = qs.filter(
            Q(descripcion__icontains=q) |
            Q(referencia__icontains=q) |
            Q(banco__icontains=q) |
            Q(cuenta__icontains=q)
        )

    desde_raw = request.GET.get("desde")
    hasta_raw = request.GET.get("hasta")
    desde = parse_date(desde_raw) if desde_raw else None
    hasta = parse_date(hasta_raw) if hasta_raw else None

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)

    solo_no = (request.GET.get("solo_no_conciliados") or "") == "1"
    if solo_no:
        qs = qs.filter(conciliado=False)

    total = qs.aggregate(
        total=Coalesce(
            Sum('monto'),
            Value(Decimal('0.00')),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    )['total']

    total_movs = qs.count()
    total_no = qs.filter(conciliado=False).count()
    total_si = qs.filter(conciliado=True).count()

    movimientos = qs.order_by("-fecha", "-id")[:500]

    return render(request, "pagos/cartolas_lista.html", {
        "movimientos": movimientos,
        "total": total,
        "q": q,
        "desde": desde,
        "hasta": hasta,
        "solo_no_conciliados": solo_no,
        "kpi_total": total_movs,
        "kpi_conciliados": total_si,
        "kpi_pendientes": total_no,
    })


# ==================================================
# SUGERENCIAS DE CONCILIACIÓN (MATCHING)
# ==================================================

@staff_member_required
def cartolas_sugerencias(request):
    mov_id = request.GET.get("mov_id")

    if mov_id:
        mov = get_object_or_404(MovimientoBancario, id=mov_id)

        if getattr(mov, "conciliado", False):
            messages.info(request, "Este movimiento ya está conciliado.")
            return redirect("cartolas_lista")

        movs = [mov]
        limite = 1
        desde = None
        hasta = None
        solo_pendientes = True
    else:
        limite = int(request.GET.get("limite") or 100)
        if limite < 10:
            limite = 10
        if limite > 500:
            limite = 500

        desde_raw = request.GET.get("desde")
        hasta_raw = request.GET.get("hasta")
        desde = parse_date(desde_raw) if desde_raw else None
        hasta = parse_date(hasta_raw) if hasta_raw else None

        solo_pendientes = (request.GET.get("solo_pendientes") or "1") == "1"

        mov_qs = MovimientoBancario.objects.all()
        if desde:
            mov_qs = mov_qs.filter(fecha__gte=desde)
        if hasta:
            mov_qs = mov_qs.filter(fecha__lte=hasta)

        if solo_pendientes:
            mov_qs = mov_qs.filter(conciliado=False)

        movs = list(mov_qs.order_by("-fecha", "-id")[:limite])

    sugerencias = []
    pagos_ocupados_en_pantalla = set()

    for m in movs:
        ranked = _obtener_candidatos_para_movimiento(
            m,
            max_candidatos=25,
            excluir_pago_ids=pagos_ocupados_en_pantalla
        )

        top = ranked[:3]

        if top:
            sugerencias.append({
                "mov": m,
                "candidatos": [{"score": s, "pago": p} for s, p in top]
            })

    auto_form = AutoConciliacionForm(initial={
        "score_minimo": 85,
        "limite": limite,
        "desde": desde,
        "hasta": hasta,
        "solo_pendientes": solo_pendientes,
    })

    return render(request, "pagos/cartolas_sugerencias.html", {
        "sugerencias": sugerencias,
        "limite": limite,
        "desde": desde,
        "hasta": hasta,
        "solo_pendientes": solo_pendientes,
        "mov_id": mov_id,
        "auto_form": auto_form,
    })


# ==================================================
# AUTO-CONCILIACIÓN
# ==================================================

@staff_member_required
def cartolas_auto_conciliar(request):
    if request.method != "POST":
        return redirect("cartolas_sugerencias")

    form = AutoConciliacionForm(request.POST)
    if not form.is_valid():
        messages.error(request, "No se pudo ejecutar la auto-conciliación. Revisa los parámetros.")
        return redirect("cartolas_sugerencias")

    score_minimo = form.cleaned_data.get("score_minimo") or 85
    limite = form.cleaned_data.get("limite") or 200
    desde = form.cleaned_data.get("desde")
    hasta = form.cleaned_data.get("hasta")
    solo_pendientes = form.cleaned_data.get("solo_pendientes")

    qs = MovimientoBancario.objects.all()

    if desde:
        qs = qs.filter(fecha__gte=desde)
    if hasta:
        qs = qs.filter(fecha__lte=hasta)

    if solo_pendientes:
        qs = qs.filter(conciliado=False)

    qs = qs.filter(tipo="cargo").order_by("-fecha", "-id")[:limite]

    conciliados = 0
    sin_match = 0
    conflictos = 0
    omitidos = 0

    pagos_reservados = set(
        MovimientoBancario.objects.filter(
            conciliado=True,
            pago_real__isnull=False
        ).values_list("pago_real_id", flat=True)
    )

    for mov in qs:
        if mov.conciliado:
            omitidos += 1
            continue

        ranked = _obtener_candidatos_para_movimiento(
            mov,
            max_candidatos=25,
            excluir_pago_ids=pagos_reservados
        )

        if not ranked:
            sin_match += 1
            continue

        top_score, top_pago = ranked[0]

        if top_score < score_minimo:
            sin_match += 1
            continue

        if len(ranked) > 1:
            segundo_score = ranked[1][0]
            if (top_score - segundo_score) <= 3:
                conflictos += 1
                continue

        if _pago_ya_conciliado_con_otro_movimiento(top_pago.id, mov_id=mov.id):
            conflictos += 1
            continue

        mov.marcar_conciliado(
            top_pago,
            nota=f"Auto-conciliado (score {top_score})"
        )
        pagos_reservados.add(top_pago.id)
        conciliados += 1

    messages.success(
        request,
        f"Auto-conciliación finalizada ✅ "
        f"Conciliados: {conciliados} | "
        f"Omitidos: {omitidos} | "
        f"Sin match fuerte: {sin_match} | "
        f"Conflictos: {conflictos}"
    )

    params = []
    if desde:
        params.append(f"desde={desde.strftime('%Y-%m-%d')}")
    if hasta:
        params.append(f"hasta={hasta.strftime('%Y-%m-%d')}")
    params.append(f"limite={limite}")
    params.append(f"solo_pendientes={'1' if solo_pendientes else '0'}")

    url = "/pagos/cartolas/sugerencias/"
    if params:
        url += "?" + "&".join(params)
    return redirect(url)


# ==================================================
# PANEL DE CONCILIACIÓN
# ==================================================

@staff_member_required
def conciliacion_panel(request):
    movimientos = MovimientoBancario.objects.all()
    pagos = PagoReal.objects.select_related("pago").all()

    total_movimientos = movimientos.count()
    total_conciliados = movimientos.filter(conciliado=True).count()
    total_pendientes = movimientos.filter(conciliado=False).count()

    porcentaje_conciliacion = 0
    if total_movimientos > 0:
        porcentaje_conciliacion = round((total_conciliados / total_movimientos) * 100, 2)

    pagos_con_movimiento_ids = set(
        movimientos.filter(
            conciliado=True,
            pago_real__isnull=False
        ).values_list("pago_real_id", flat=True)
    )

    pagos_sin_movimiento = pagos.exclude(id__in=pagos_con_movimiento_ids)

    monto_conciliado = movimientos.filter(conciliado=True).aggregate(
        total=Coalesce(
            Sum("monto"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    )["total"]

    monto_pendiente = movimientos.filter(conciliado=False).aggregate(
        total=Coalesce(
            Sum("monto"),
            Value(Decimal("0.00")),
            output_field=DecimalField(max_digits=14, decimal_places=2)
        )
    )["total"]

    conciliados_recientes = (
        movimientos
        .filter(conciliado=True)
        .select_related("pago_real", "pago_real__pago")
        .order_by("-conciliado_en", "-id")[:10]
    )

    pendientes_recientes = (
        movimientos
        .filter(conciliado=False)
        .select_related("pago_real", "pago_real__pago")
        .order_by("-fecha", "-id")[:10]
    )

    return render(request, "pagos/conciliacion_panel.html", {
        "total_movimientos": total_movimientos,
        "total_conciliados": total_conciliados,
        "total_pendientes": total_pendientes,
        "porcentaje_conciliacion": porcentaje_conciliacion,
        "total_pagos_reales": pagos.count(),
        "pagos_sin_movimiento_count": pagos_sin_movimiento.count(),
        "pagos_sin_movimiento": pagos_sin_movimiento.order_by("-fecha_pago", "-id")[:10],
        "monto_conciliado": monto_conciliado,
        "monto_pendiente": monto_pendiente,
        "conciliados_recientes": conciliados_recientes,
        "pendientes_recientes": pendientes_recientes,
    })


# ==================================================
# ACCIONES: CONCILIAR / DESCONCILIAR
# ==================================================

@staff_member_required
def cartolas_conciliar(request):
    if request.method != "POST":
        return redirect("cartolas_sugerencias")

    mov_id = request.POST.get("mov_id")
    pago_real_id = request.POST.get("pago_real_id")
    nota = (request.POST.get("nota") or "").strip()

    if not mov_id or not pago_real_id:
        messages.error(request, "Faltan parámetros para conciliar.")
        return redirect("cartolas_sugerencias")

    mov = get_object_or_404(MovimientoBancario, id=mov_id)
    pago = get_object_or_404(PagoReal.objects.select_related("pago"), id=pago_real_id)

    if getattr(mov, "conciliado", False) and getattr(mov, "pago_real_id", None) == pago.id:
        messages.info(request, "Este movimiento ya estaba conciliado con ese pago.")
        return redirect("cartolas_lista")

    if _pago_ya_conciliado_con_otro_movimiento(pago.id, mov_id=mov.id):
        messages.error(
            request,
            "Ese pago real ya está conciliado con otro movimiento bancario. "
            "Primero debes desconciliar el anterior si corresponde."
        )
        return redirect("cartolas_sugerencias")

    warn = []
    if mov.monto != pago.monto:
        diff = abs(mov.monto - pago.monto)
        if diff > Decimal("50.00"):
            warn.append(f"Monto difiere en ${diff:,.0f}")

    if pago.fecha_pago:
        delta = abs((mov.fecha - pago.fecha_pago).days)
        if delta > 7:
            warn.append(f"Fecha difiere en {delta} día(s)")

    if hasattr(mov, "marcar_conciliado"):
        mov.marcar_conciliado(pago, nota=nota)
    else:
        mov.pago_real = pago
        mov.conciliado = True
        mov.conciliado_en = timezone.now()
        mov.nota_conciliacion = (nota or "")[:255]
        mov.save()

    if warn:
        messages.warning(request, "Conciliado ✅ (ojo: " + " | ".join(warn) + ")")
    else:
        messages.success(request, "Conciliado ✅")

    next_url = request.POST.get("next") or "cartolas_lista"
    return redirect(next_url)


@staff_member_required
def cartolas_desconciliar(request, mov_id: int):
    mov = get_object_or_404(MovimientoBancario, id=mov_id)

    if getattr(mov, "conciliado", False):
        if hasattr(mov, "desconciliar"):
            mov.desconciliar()
        else:
            mov.pago_real = None
            mov.conciliado = False
            mov.conciliado_en = None
            mov.nota_conciliacion = ""
            mov.save()
        messages.success(request, "Movimiento desconciliado ✅")
    else:
        messages.info(request, "Este movimiento ya estaba sin conciliar.")

    return redirect("cartolas_lista")


# ==================================================
# DESCARGAS IMPORTACIÓN
# ==================================================

@staff_member_required
def descargar_plantilla_importacion(request):
    try:
        file_path = os.path.join(
            settings.BASE_DIR,
            'pagos',
            'static',
            'pagos',
            'docs',
            'plantilla_importacion_pagos.xlsx'
        )

        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename='plantilla_importacion_pagos.xlsx'
        )
    except Exception as e:
        messages.error(request, f"No se encontró la plantilla Excel. ({e})")
        return redirect('pagos_importar_excel')


@staff_member_required
def descargar_guia_importacion(request):
    try:
        file_path = os.path.join(
            settings.BASE_DIR,
            'pagos',
            'static',
            'pagos',
            'docs',
            'guia_importacion_pagos.pdf'
        )

        return FileResponse(
            open(file_path, 'rb'),
            as_attachment=True,
            filename='guia_importacion_pagos.pdf'
        )
    except Exception as e:
        messages.error(request, f"No se encontró la guía PDF. ({e})")
        return redirect('pagos_importar_excel')